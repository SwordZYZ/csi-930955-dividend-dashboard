import os
import io
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

# 🎯 核心路径安全锁：无论谁在哪个文件夹调用这个脚本，
# 都以当前脚本文件所在位置为基准计算，死死锁定项目根目录下的 xlsx 文件。
CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(CURRENT_DIR) # 👈 往上一层就是根目录
OUTPUT_FILE = os.path.join(PROJECT_ROOT, "930955indicator.xlsx")

URL = "https://oss-ch.csindex.com.cn/static/html/csindex/public/uploads/file/autofile/indicator/930955indicator.xls"
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
}

def fetch_csi_930955_list(limit_days: int = 20) -> list:
    """
    1. 抓取中证官网最新单日数据，并安全更新进本地 Excel 库。
    2. 读取 Excel 库，逆向捞出最近 limit_days 天的历史数据，返回升序 list 给前端全量替换。
    """
    # ------------------ 第一步：日常抓取与 Excel 插入 ------------------
    try:
        print("🚀 正在获取中证官网最新红利低波 100 指数数据...")
        response = requests.get(URL, headers=HEADERS, timeout=15)
        response.raise_for_status()
        
        df_web = pd.read_excel(io.BytesIO(response.content))
        if df_web.empty:
            print("⚠️ 未能在网上获取到有价值数据。")
        else:
            headers_list = df_web.columns.tolist()
            latest_row_values = df_web.iloc[0].tolist()
            current_date = latest_row_values[0]
            print(f"📈 成功捕获官网当日最新数据，日期为: {current_date}")

            # 写入本地 Excel 存储仓库
            if os.path.exists(OUTPUT_FILE):
                wb = load_workbook(OUTPUT_FILE)
                ws = wb.active
                
                # 查重防止重复插入
                date_exists = False
                for row in range(2, ws.max_row + 1):
                    if str(ws.cell(row=row, column=1).value) == str(current_date):
                        date_exists = True
                        break
                
                if date_exists:
                    print(f"💤 查重提示：本地 Excel 已存在 {current_date} 数据，跳过插入。")
                else:
                    ws.insert_rows(idx=2, amount=1)
                    for col_idx, value in enumerate(latest_row_values, start=1):
                        ws.cell(row=2, column=col_idx, value=value)
                    wb.save(OUTPUT_FILE)
                    print("💾 成功将最新单日数据追加进本地 Excel 库！")
            else:
                print("🆕 未发现本地 Excel 历史库，正在自动初始化创建...")
                wb = Workbook()
                ws = wb.active
                ws.append(headers_list)
                ws.append(latest_row_values)
                wb.save(OUTPUT_FILE)

    except Exception as e:
        # 即使官网挂了/网络抖动，我们也绝不崩溃，记录错误后继续走下一步（去 Excel 捞历史），确保自愈性！
        print(f"⚠️ 抓取官网单日数据失败 (将直接使用历史库兜底), 原因: {e}")

    # ------------------ 第二步：从本地 Excel 逆向输出 List 给前端 ------------------
    try:
        if not os.path.exists(OUTPUT_FILE):
            print("❌ 致命错误: 找不到本地历史 Excel 数据库文件，无法提供前端列表！")
            return []

        # 用 pandas 读取刚才更新完的本地库
        df_local = pd.read_excel(OUTPUT_FILE)
        
        # 提取最近的 20 天历史行（注意：你的 Excel 库里第二行是最新的，所以前 20 行就是最近 20 天）
        df_recent = df_local.head(limit_days).copy()

        cleaned_list = []
        for _, row in df_recent.iterrows():

            cleaned_list.append({
                "日期Date": int(row['日期Date']),
                "股息率1（总股本）D/P1": float(row['股息率1（总股本）D/P1']),
                "股息率2（计算用股本）D/P2": float(row['股息率2（计算用股本）D/P2'])
            })

        # 🎯 灵魂反转：Excel 里上面最新、下面最旧。
        # 翻转后，变成旧数据在前，最新数据在后，完美符合 ECharts 渲染顺序
        cleaned_list.reverse()
        
        print(f"✅ 红利低波组件输出成功，共汇编 {len(cleaned_list)} 个最新历史交易日 List。")
        return cleaned_list

    except Exception as e:
        print(f"❌ 解析本地历史库生成列表失败: {e}")
        return []

# 💡 留给你的“单兵训练”调试接口
# 无论你在根目录下敲 `uv run fetchers/csi_indicator.py` 
# 还是 cd 到 fetchers 目录敲 `uv run csi_indicator.py`，因为有了上面的路径锁，都能完美独立运行调试！
if __name__ == "__main__":
    res = fetch_csi_930955_list(limit_days=20) # 测试抓取最近 20 天
    print("\n[本地调试输出样例（前两项）]:")
    import json
    print(json.dumps(res[:2], ensure_ascii=False, indent=2))